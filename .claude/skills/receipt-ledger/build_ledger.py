#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
build_ledger.py — 영수증 회계장부 스킬의 실행 스크립트

역할 (에이전트가 영수증을 읽어 만든 JSON을 입력으로 받아):
  1) 원본을 보존하면서 영수증 사본을 Output/<카테고리>/ 로 정리
  2) 회계장부_실습.xlsx 양식(영수증입력 / 회계장부 / 카테고리별집계 3시트)을
     그대로 따르는 결과 엑셀을 Output/ 안에 생성

사용법:
  python build_ledger.py <input.json>

input.json 형식:
{
  "source_dir":   "<원본 영수증 이미지 폴더>",
  "output_dir":   "<결과 Output 폴더>",
  "output_name":  "회계장부_결과.xlsx",   # 선택, 기본값 동일
  "opening_balance": 500000,               # 선택, 기본 500000 (전기이월/기초잔액)
  "opening_date": "2026-06-01",            # 선택, 없으면 가장 이른 영수증 날짜
  "receipts": [
    {
      "file": "receipt_ko_01.png",   # source_dir 기준 파일명 (사본 정리용, 없으면 스킵)
      "date": "2026-06-02",          # YYYY-MM-DD
      "merchant": "하늘카페 강남점",   # 가맹점명
      "amount": 28100,               # 합계 금액 (정수)
      "payment": "현금",              # 결제수단
      "category": "식비",             # 식비/교통비/도서비/사무용품/기타
      "memo": "카페라떼 외",          # 비고 (품목 요약) - 선택
      "desc": "카페·디저트"           # 회계장부 적요 보조설명 - 선택
    }
  ]
}
"""
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---- 회계장부_실습.xlsx 양식과 동일한 서식 상수 -------------------------------
NUMFMT = '#,##0;(#,##0);"-"'          # 금액 표시 형식
PCTFMT = '0.0%'
C_HEADER = "2F5496"                   # 헤더 배경 (진한 파랑)
C_TITLE = "2F5496"                    # 제목 글자색
C_ALT = "F2F7FB"                      # 짝수행 옅은 배경
C_TOTAL = "D6E4F0"                    # 합계행 배경
C_BORDER = "BFBFBF"                   # 얇은 테두리색

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill("solid", fgColor=C_HEADER)
FILL_ALT = PatternFill("solid", fgColor=C_ALT)
FILL_TOTAL = PatternFill("solid", fgColor=C_TOTAL)

# 카테고리 표시 순서 (양식 기준). 실제 데이터에 있는 것만 집계에 표기한다.
CATEGORY_ORDER = ["식비", "교통비", "도서비", "사무용품", "기타"]

INVALID_CHARS = re.compile(r'[\\/:*?"<>|]')


def _sanitize(name: str) -> str:
    """Windows 파일명으로 안전하게."""
    return INVALID_CHARS.sub("_", name).strip().rstrip(".")


def _title(ws, text, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, size=14, color=C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 4  # 제목과 표 사이 여백행


def _header(ws, headers, row=3):
    ws.row_dimensions[row].height = 22
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def _style_cell(c, *, align="center", numfmt=None, alt=False,
                bold=False, total=False):
    c.font = Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if numfmt:
        c.number_format = numfmt
    if total:
        c.fill = FILL_TOTAL
    elif alt:
        c.fill = FILL_ALT


def _set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build(data: dict) -> Path:
    source_dir = Path(data["source_dir"])
    output_dir = Path(data["output_dir"])
    output_name = data.get("output_name", "회계장부_결과.xlsx")
    opening_balance = int(data.get("opening_balance", 500000))
    receipts = list(data["receipts"])

    if not receipts:
        raise SystemExit("영수증 데이터가 비어 있습니다.")

    # 날짜 → 파일명 순으로 정렬 (안정적 순서)
    receipts.sort(key=lambda r: (r.get("date", ""), r.get("file", "")))
    opening_date = data.get("opening_date") or receipts[0].get("date", "")

    output_dir.mkdir(parents=True, exist_ok=True)

    # ---- 1) 카테고리별 폴더로 사본 정리 (원본 보존) ------------------------
    copied = []
    for r in receipts:
        fname = r.get("file")
        cat = r.get("category", "기타")
        if not fname:
            continue
        src = source_dir / fname
        if not src.exists():
            print(f"  [경고] 원본 없음, 건너뜀: {src}")
            continue
        cat_dir = output_dir / cat
        cat_dir.mkdir(parents=True, exist_ok=True)
        newname = f"{r.get('date','')}_{_sanitize(r.get('merchant',''))}{src.suffix}"
        dst = cat_dir / newname
        # 이름 충돌 방지
        i = 2
        while dst.exists():
            dst = cat_dir / f"{r.get('date','')}_{_sanitize(r.get('merchant',''))}_{i}{src.suffix}"
            i += 1
        shutil.copy2(src, dst)
        copied.append((cat, dst.name))

    # ---- 2) 엑셀 생성 (양식 3시트) ---------------------------------------
    wb = openpyxl.Workbook()

    # === 시트1: 영수증입력 =================================================
    ws1 = wb.active
    ws1.title = "영수증입력"
    _set_widths(ws1, {"A": 13, "B": 20, "C": 13, "D": 13, "E": 13, "F": 18})
    _title(ws1, "영수증 입력 장부", 6)
    _header(ws1, ["날짜", "가맹점명", "금액", "결제수단", "카테고리", "비고"])
    row = 4
    for idx, r in enumerate(receipts):
        alt = idx % 2 == 1
        ws1.row_dimensions[row].height = 18
        vals = [
            (r.get("date", ""), "center", None),
            (r.get("merchant", ""), "left", None),
            (int(r.get("amount", 0)), "center", NUMFMT),
            (r.get("payment", ""), "center", None),
            (r.get("category", ""), "center", None),
            (r.get("memo", ""), "center", None),
        ]
        for col, (val, align, nf) in enumerate(vals, start=1):
            c = ws1.cell(row=row, column=col, value=val)
            _style_cell(c, align=align, numfmt=nf, alt=alt)
        row += 1

    # === 시트2: 회계장부 ===================================================
    ws2 = wb.create_sheet("회계장부")
    _set_widths(ws2, {"A": 13, "B": 28, "C": 14, "D": 14, "E": 14, "F": 13})
    _title(ws2, "회  계  장  부", 6)
    _header(ws2, ["날짜", "적요 (거래 내용)", "차변 (지출)",
                  "대변 (수입)", "잔  액", "계정과목"])
    # 전기이월(기초잔액) 행
    ws2.row_dimensions[4].height = 18
    _style_cell(ws2.cell(row=4, column=1, value=opening_date), align="center")
    _style_cell(ws2.cell(row=4, column=2, value="전기이월 (기초잔액)"), align="left")
    _style_cell(ws2.cell(row=4, column=3), numfmt=NUMFMT)
    _style_cell(ws2.cell(row=4, column=4, value=opening_balance), numfmt=NUMFMT)
    _style_cell(ws2.cell(row=4, column=5, value="=D4"), numfmt=NUMFMT)
    _style_cell(ws2.cell(row=4, column=6, value="현금·예금"), align="center")
    # 지출 행
    row = 5
    for idx, r in enumerate(receipts):
        alt = idx % 2 == 0  # 양식: 전기이월 다음 첫 지출행부터 옅은배경 시작
        ws2.row_dimensions[row].height = 18
        desc = r.get("desc", "")
        jeokyo = f"{r.get('merchant','')} {desc}".strip()
        _style_cell(ws2.cell(row=row, column=1, value=r.get("date", "")),
                    align="center", alt=alt)
        _style_cell(ws2.cell(row=row, column=2, value=jeokyo),
                    align="left", alt=alt)
        _style_cell(ws2.cell(row=row, column=3, value=int(r.get("amount", 0))),
                    numfmt=NUMFMT, alt=alt)
        _style_cell(ws2.cell(row=row, column=4), numfmt=NUMFMT, alt=alt)
        bal = (f"=E{row-1}-IF(ISNUMBER(C{row}),C{row},0)"
               f"+IF(ISNUMBER(D{row}),D{row},0)")
        _style_cell(ws2.cell(row=row, column=5, value=bal), numfmt=NUMFMT, alt=alt)
        _style_cell(ws2.cell(row=row, column=6, value=r.get("category", "")),
                    align="center", alt=alt)
        row += 1
    last = row - 1
    # 합계 행
    ws2.row_dimensions[row].height = 18
    _style_cell(ws2.cell(row=row, column=1), total=True)
    _style_cell(ws2.cell(row=row, column=2, value="합  계"), align="left",
                bold=True, total=True)
    _style_cell(ws2.cell(row=row, column=3, value=f"=SUM(C5:C{last})"),
                numfmt=NUMFMT, bold=True, total=True)
    _style_cell(ws2.cell(row=row, column=4, value=f"=SUM(D4:D{last})"),
                numfmt=NUMFMT, bold=True, total=True)
    _style_cell(ws2.cell(row=row, column=5, value=f"=E{last}"),
                numfmt=NUMFMT, bold=True, total=True)
    _style_cell(ws2.cell(row=row, column=6), total=True)

    # === 시트3: 카테고리별집계 ============================================
    ws3 = wb.create_sheet("카테고리별집계")
    _set_widths(ws3, {"A": 16, "B": 16, "C": 13})
    _title(ws3, "카테고리별 지출 집계", 3)
    _header(ws3, ["카테고리", "지출 합계", "비율 (%)"])
    used = [c for c in receipts]
    present = [cat for cat in CATEGORY_ORDER
              if any(r.get("category") == cat for r in used)]
    # 표준 순서에 없는 카테고리도 뒤에 추가
    for r in used:
        if r.get("category") not in present and r.get("category"):
            present.append(r.get("category"))
    row = 4
    first = row
    for idx, cat in enumerate(present):
        alt = idx % 2 == 1
        ws3.row_dimensions[row].height = 18
        _style_cell(ws3.cell(row=row, column=1, value=cat), align="center", alt=alt)
        row += 1
    last = row - 1
    total_row = row
    # SUMIF/비율 수식은 합계행 위치(total_row)를 참조
    for r in range(first, last + 1):
        _style_cell(ws3.cell(row=r, column=2,
                    value=f"=SUMIF(영수증입력!E:E,A{r},영수증입력!C:C)"),
                    numfmt=NUMFMT, alt=(r - first) % 2 == 1)
        _style_cell(ws3.cell(row=r, column=3,
                    value=f"=IF(B{total_row}=0,0,B{r}/B{total_row})"),
                    numfmt=PCTFMT, alt=(r - first) % 2 == 1)
    # 합계 행
    ws3.row_dimensions[total_row].height = 18
    _style_cell(ws3.cell(row=total_row, column=1, value="합  계"),
                align="center", bold=True, total=True)
    _style_cell(ws3.cell(row=total_row, column=2, value=f"=SUM(B{first}:B{last})"),
                numfmt=NUMFMT, bold=True, total=True)
    _style_cell(ws3.cell(row=total_row, column=3,
                value=f"=IF(B{total_row}=0,0,1)"),
                numfmt=PCTFMT, bold=True, total=True)

    out_path = output_dir / output_name
    wb.save(out_path)

    # ---- 요약 출력 -------------------------------------------------------
    print(f"[완료] 영수증 {len(receipts)}건 처리")
    by_cat = {}
    for cat, name in copied:
        by_cat.setdefault(cat, []).append(name)
    for cat in present:
        files = by_cat.get(cat, [])
        print(f"  - {cat}: {len(files)}건  →  Output/{cat}/")
    total_amount = sum(int(r.get("amount", 0)) for r in receipts)
    print(f"  지출 합계: {total_amount:,}원  |  기초잔액 {opening_balance:,}원 "
          f"→ 잔액 {opening_balance - total_amount:,}원")
    print(f"[엑셀] {out_path}")
    return out_path


def main():
    if len(sys.argv) != 2:
        print("사용법: python build_ledger.py <input.json>")
        raise SystemExit(1)
    data = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    build(data)


if __name__ == "__main__":
    main()
